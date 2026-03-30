"""
SBOM Excel 보고서 생성기
시트 구성: 요약 대시보드 (차트 포함), 컴포넌트 목록, 취약점 목록
v2 데이터 구조 (dict 기반) 호환
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference


# ─────────────────────────────────────────────
# 스타일 상수
# ─────────────────────────────────────────────
FONT_NAME = "Arial"
C_DARK = "0F172A"
C_PRIMARY = "1E40AF"
C_WHITE = "FFFFFF"
C_LIGHT_BG = "F1F5F9"
C_BORDER = "CBD5E1"
C_CRITICAL = "DC2626"
C_HIGH = "EA580C"
C_MEDIUM = "D97706"
C_LOW = "16A34A"
C_NONE = "64748B"

THIN_BORDER = Border(
    left=Side("thin", C_BORDER), right=Side("thin", C_BORDER),
    top=Side("thin", C_BORDER), bottom=Side("thin", C_BORDER),
)
HEADER_FONT = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=11)
HEADER_FILL = PatternFill("solid", fgColor=C_PRIMARY)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name=FONT_NAME, size=10)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
DATA_CENTER = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor=C_LIGHT_BG)

SEV_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="FEE2E2"),
    "HIGH": PatternFill("solid", fgColor="FFEDD5"),
    "MEDIUM": PatternFill("solid", fgColor="FEF9C3"),
    "LOW": PatternFill("solid", fgColor="DCFCE7"),
    "UNKNOWN": PatternFill("solid", fgColor=C_LIGHT_BG),
}
SEV_FONTS = {
    "CRITICAL": Font(name=FONT_NAME, bold=True, color=C_CRITICAL, size=10),
    "HIGH": Font(name=FONT_NAME, bold=True, color=C_HIGH, size=10),
    "MEDIUM": Font(name=FONT_NAME, bold=True, color=C_MEDIUM, size=10),
    "LOW": Font(name=FONT_NAME, bold=True, color=C_LOW, size=10),
    "UNKNOWN": Font(name=FONT_NAME, color=C_NONE, size=10),
}


def _hdr(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = w


def _cell(ws, row, col, value, center=False, font=None, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or DATA_FONT
    c.alignment = DATA_CENTER if center else DATA_ALIGN
    c.border = THIN_BORDER
    if fill:
        c.fill = fill
    elif row % 2 == 0:
        c.fill = ALT_FILL
    return c


def _info_row(ws, row, label, value):
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = Font(name=FONT_NAME, bold=True, size=10)
    c1.border = THIN_BORDER
    c2 = ws.cell(row=row, column=2, value=str(value))
    c2.font = DATA_FONT
    c2.border = THIN_BORDER


def _section_title(ws, row, text, merge_to="F"):
    ws.merge_cells(f"A{row}:{merge_to}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name=FONT_NAME, bold=True, size=13, color=C_PRIMARY)


# ─────────────────────────────────────────────
# 메인 함수
# ─────────────────────────────────────────────

def generate_excel(
    sbom_components: list,
    cve_vulns: list = None,
    metadata: dict = None,
) -> bytes:
    """
    SBOM + CVE 결과를 전문적인 Excel 보고서로 생성합니다.

    Args:
        sbom_components: parse_sbom_components() 결과 (list of dict)
        cve_vulns: vulns_to_table() 결과 (list of dict)
        metadata: {"tool", "scan_path", "input_type", "elapsed", ...}

    Returns:
        bytes (.xlsx)
    """
    wb = Workbook()
    meta = metadata or {}
    vulns = cve_vulns or []

    ws_summary = wb.active
    ws_summary.title = "요약 대시보드"
    ws_comp = wb.create_sheet("컴포넌트 목록")
    ws_vuln = None
    if vulns:
        ws_vuln = wb.create_sheet("취약점 목록")

    _build_summary(ws_summary, sbom_components, vulns, meta)
    _build_components(ws_comp, sbom_components)
    if ws_vuln:
        _build_vulns(ws_vuln, vulns)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# 시트 1: 요약 대시보드
# ─────────────────────────────────────────────

def _build_summary(ws, components, vulns, meta):
    ws.sheet_properties.tabColor = C_PRIMARY

    # 제목
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "SBOM & Vulnerability Analysis Report"
    t.font = Font(name=FONT_NAME, bold=True, size=18, color=C_DARK)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"].value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
        f"Tool: {meta.get('tool', 'SBOM Generator')}  |  v2.0.0"
    )
    ws["A2"].font = Font(name=FONT_NAME, size=10, color=C_NONE)
    ws.row_dimensions[2].height = 22

    # ── 프로젝트 정보 ──
    row = 4
    _section_title(ws, row, "프로젝트 정보")
    row += 1

    info_items = [
        ("스캔 경로", meta.get("scan_path", "-")),
        ("입력 유형", meta.get("input_type", "-")),
        ("SBOM 도구", meta.get("tool", "-")),
        ("CVE 도구", meta.get("cve_tools", "-")),
        ("소요 시간", f"{meta.get('elapsed', 0):.1f}초"),
        ("총 컴포넌트 수", len(components)),
        ("총 취약점 수", len(vulns)),
    ]
    for label, value in info_items:
        _info_row(ws, row, label, value)
        row += 1

    # ── 생태계별 분포 ──
    row += 1
    _section_title(ws, row, "생태계별 컴포넌트 분포")
    row += 1

    eco_counts = {}
    for c in components:
        eco = c.get("ecosystem", "-")
        eco_counts[eco] = eco_counts.get(eco, 0) + 1

    eco_header_row = row
    _hdr(ws, row, ["생태계", "컴포넌트 수", "비율"], [30, 15, 15])
    row += 1

    eco_data_start = row
    total = len(components) or 1
    for eco, cnt in sorted(eco_counts.items(), key=lambda x: -x[1]):
        _cell(ws, row, 1, eco)
        _cell(ws, row, 2, cnt, center=True)
        pct_cell = _cell(ws, row, 3, cnt / total, center=True)
        pct_cell.number_format = "0.0%"
        row += 1

    # 파이 차트
    if len(eco_counts) > 1:
        pie = PieChart()
        pie.title = "생태계 분포"
        pie.style = 10
        pie.width = 16
        pie.height = 10
        labels = Reference(ws, min_col=1, min_row=eco_data_start, max_row=row - 1)
        data = Reference(ws, min_col=2, min_row=eco_header_row, max_row=row - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, f"D{eco_header_row}")

    # ── 취약점 요약 ──
    if vulns:
        row += 8  # 차트 공간
        _section_title(ws, row, "취약점 분석 요약")
        row += 1

        # 심각도별 카운트
        sev_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0, "UNKNOWN": 0}
        for v in vulns:
            sev = v.get("심각도", "UNKNOWN")
            sev_counts[sev] = sev_counts.get(sev, 0) + 1

        vuln_info = [
            ("총 취약점 수", len(vulns)),
            ("CRITICAL", sev_counts["CRITICAL"]),
            ("HIGH", sev_counts["HIGH"]),
            ("MEDIUM", sev_counts["MEDIUM"]),
            ("LOW", sev_counts["LOW"]),
        ]
        for label, value in vuln_info:
            _info_row(ws, row, label, value)
            row += 1

        # 심각도 바 차트
        row += 1
        sev_hdr_row = row
        _hdr(ws, row, ["심각도", "건수", "비율"], [20, 12, 12])
        row += 1

        vuln_total = len(vulns) or 1
        sev_data = [
            ("CRITICAL", sev_counts["CRITICAL"], C_CRITICAL),
            ("HIGH", sev_counts["HIGH"], C_HIGH),
            ("MEDIUM", sev_counts["MEDIUM"], C_MEDIUM),
            ("LOW", sev_counts["LOW"], C_LOW),
        ]
        for sev_name, cnt, color in sev_data:
            sfont = Font(name=FONT_NAME, bold=True, color=color, size=10)
            _cell(ws, row, 1, sev_name, font=sfont)
            _cell(ws, row, 2, cnt, center=True)
            p = _cell(ws, row, 3, cnt / vuln_total, center=True)
            p.number_format = "0.0%"
            row += 1

        bar = BarChart()
        bar.type = "col"
        bar.title = "심각도별 취약점 분포"
        bar.style = 10
        bar.width = 16
        bar.height = 10
        bar.y_axis.title = "건수"
        d_ref = Reference(ws, min_col=2, min_row=sev_hdr_row, max_row=row - 1)
        c_ref = Reference(ws, min_col=1, min_row=sev_hdr_row + 1, max_row=row - 1)
        bar.add_data(d_ref, titles_from_data=True)
        bar.set_categories(c_ref)
        bar.shape = 4
        ws.add_chart(bar, f"D{sev_hdr_row}")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50


# ─────────────────────────────────────────────
# 시트 2: 컴포넌트 목록
# ─────────────────────────────────────────────

def _build_components(ws, components):
    ws.sheet_properties.tabColor = "1E40AF"

    headers = ["No.", "생태계", "패키지명", "그룹", "버전", "유형", "라이선스", "PURL"]
    widths = [6, 16, 35, 20, 15, 12, 20, 55]
    _hdr(ws, 1, headers, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(len(components) + 1, 2)}"

    for idx, comp in enumerate(components, 1):
        r = idx + 1
        _cell(ws, r, 1, idx, center=True)
        _cell(ws, r, 2, comp.get("ecosystem", ""))
        _cell(ws, r, 3, comp.get("name", ""))
        _cell(ws, r, 4, comp.get("group", ""))
        _cell(ws, r, 5, comp.get("version", ""), center=True)
        _cell(ws, r, 6, comp.get("type", ""))
        _cell(ws, r, 7, comp.get("licenses", ""))
        _cell(ws, r, 8, comp.get("purl", ""))


# ─────────────────────────────────────────────
# 시트 3: 취약점 목록
# ─────────────────────────────────────────────

def _build_vulns(ws, vulns):
    ws.sheet_properties.tabColor = "DC2626"

    headers = [
        "No.", "심각도", "CVSS", "CVE ID", "CWE",
        "패키지명", "현재 버전", "수정 버전", "출처",
        "설명", "참조 링크"
    ]
    widths = [6, 12, 10, 20, 18, 30, 12, 12, 14, 55, 40]
    _hdr(ws, 1, headers, widths)
    ws.freeze_panes = "A2"
    total_rows = len(vulns)
    ws.auto_filter.ref = f"A1:K{max(total_rows + 1, 2)}"

    for idx, v in enumerate(vulns, 1):
        r = idx + 1
        sev = v.get("심각도", "UNKNOWN")
        sf = SEV_FILLS.get(sev)
        sfont = SEV_FONTS.get(sev, DATA_FONT)

        _cell(ws, r, 1, idx, center=True)
        _cell(ws, r, 2, sev, center=True, font=sfont, fill=sf)
        cvss_c = _cell(ws, r, 3, v.get("CVSS", 0), center=True, fill=sf)
        cvss_c.number_format = "0.0"
        _cell(ws, r, 4, v.get("CVE ID", ""))
        _cell(ws, r, 5, v.get("CWE", ""))
        _cell(ws, r, 6, v.get("패키지", ""))
        _cell(ws, r, 7, v.get("현재 버전", ""), center=True)

        fx = v.get("수정 버전", "-")
        fx_font = Font(name=FONT_NAME, bold=True, color=C_LOW, size=10) if fx != "-" else DATA_FONT
        _cell(ws, r, 8, fx, center=True, font=fx_font)

        _cell(ws, r, 9, v.get("출처", ""), center=True)
        _cell(ws, r, 10, v.get("설명", ""))
        _cell(ws, r, 11, v.get("참조", ""))

    # 하단 합계
    sum_row = total_rows + 3
    ws.cell(row=sum_row, column=1, value="합계").font = Font(name=FONT_NAME, bold=True, size=10)
    ws.cell(row=sum_row, column=2, value=f"=COUNTA(B2:B{total_rows+1})").font = Font(name=FONT_NAME, bold=True, size=10)

    # 심각도별 카운트
    sum_row += 2
    ws.cell(row=sum_row, column=1, value="심각도별 합계").font = Font(name=FONT_NAME, bold=True, size=11, color=C_PRIMARY)

    sev_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for v in vulns:
        sev = v.get("심각도", "")
        if sev in sev_counts:
            sev_counts[sev] += 1

    for i, (sev, cnt) in enumerate(sev_counts.items()):
        r = sum_row + 1 + i
        ws.cell(row=r, column=1, value=sev).font = SEV_FONTS.get(sev, DATA_FONT)
        ws.cell(row=r, column=2, value=cnt).font = Font(name=FONT_NAME, bold=True, size=10)
        ws.cell(row=r, column=1).fill = SEV_FILLS.get(sev, PatternFill())
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).border = THIN_BORDER
